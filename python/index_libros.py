#!/usr/bin/env python3
"""
index_libros.py
Recorre una carpeta (y subcarpetas) buscando archivos de libros
y genera un Excel con nombre, extensión, tamaño y ruta.

Uso:
    python index_libros.py /ruta/a/tu/carpeta
    python index_libros.py /ruta/a/tu/carpeta --output mi_biblioteca.xlsx

    python index_libros.py "D:\biblioteca"
    python index_libros.py "D:\biblioteca" --output "D:\biblioteca\indice.xlsx"
"""

import os
import sys
import argparse
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando openpyxl...")
    os.system(f'"{sys.executable}" -m pip install openpyxl -q')
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


EXTENSIONES = {".pdf", ".epub", ".mobi", ".azw", ".azw3", ".azw4", ".lit", ".djvu", ".cbz", ".cbr"}

COLORES_EXT = {
    ".pdf":   "FFD6E4BC",
    ".epub":  "FFB8D4E8",
    ".mobi":  "FFFCE4D6",
    ".azw":   "FFE8D5F5",
    ".azw3":  "FFE8D5F5",
    ".azw4":  "FFE8D5F5",
    ".lit":   "FFFFF2CC",
    ".djvu":  "FFDDEBF7",
    ".cbz":   "FFFFD7D7",
    ".cbr":   "FFFFD7D7",
}

COLOR_HEADER = "FF2F4F8F"
COLOR_FILA_PAR = "FFF5F5F5"


def formato_tamanio(bytes_: int) -> str:
    for unidad in ("B", "KB", "MB", "GB"):
        if bytes_ < 1024:
            return f"{bytes_:.1f} {unidad}"
        bytes_ /= 1024
    return f"{bytes_:.1f} GB"


def recorrer_carpeta(raiz: Path) -> list[dict]:
    registros = []
    for archivo in sorted(raiz.rglob("*")):
        if archivo.is_file() and archivo.suffix.lower() in EXTENSIONES:
            stat = archivo.stat()
            registros.append({
                "nombre":    archivo.stem,
                "extension": archivo.suffix.lower(),
                "tamanio_bytes": stat.st_size,
                "tamanio":   formato_tamanio(stat.st_size),
                "carpeta":   str(archivo.parent.relative_to(raiz)),
                "ruta":      str(archivo),
            })
    return registros


def thin_border():
    thin = Side(style="thin", color="FFD0D0D0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def crear_excel(registros: list[dict], output_path: Path, carpeta_raiz: Path):
    wb = Workbook()

    # ── Hoja 1: Índice completo ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Índice"

    encabezados = ["#", "Nombre", "Extensión", "Tamaño", "Carpeta", "Ruta completa"]
    anchos      = [6,   45,       12,           12,        40,         60]

    # Header
    header_font  = Font(name="Arial", bold=True, color="FFFFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color=COLOR_HEADER)
    header_align = Alignment(horizontal="center", vertical="center")

    for col, (titulo, ancho) in enumerate(zip(encabezados, anchos), 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = ancho

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Datos
    for i, reg in enumerate(registros, 1):
        fila = i + 1
        es_par = i % 2 == 0
        bg = PatternFill("solid", start_color=COLORES_EXT.get(reg["extension"], COLOR_FILA_PAR if es_par else "FFFFFFFF"))

        valores = [i, reg["nombre"], reg["extension"], reg["tamanio"], reg["carpeta"], reg["ruta"]]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=fila, column=col, value=val)
            cell.font   = Font(name="Arial", size=10)
            cell.fill   = bg
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if col in (1, 3, 4) else "left")

    # Auto-filtro
    ws.auto_filter.ref = f"A1:{get_column_letter(len(encabezados))}1"

    # ── Hoja 2: Resumen por extensión ────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")

    from collections import defaultdict
    resumen: dict[str, dict] = defaultdict(lambda: {"cantidad": 0, "bytes": 0})
    for r in registros:
        resumen[r["extension"]]["cantidad"] += 1
        resumen[r["extension"]]["bytes"]    += r["tamanio_bytes"]

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14

    enc2 = ["Extensión", "Cantidad", "Tamaño total"]
    for col, titulo in enumerate(enc2, 1):
        cell = ws2.cell(row=1, column=col, value=titulo)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border()

    for fila, (ext, datos) in enumerate(sorted(resumen.items()), 2):
        bg = PatternFill("solid", start_color=COLORES_EXT.get(ext, "FFFFFFFF"))
        for col, val in enumerate([ext, datos["cantidad"], formato_tamanio(datos["bytes"])], 1):
            cell = ws2.cell(row=fila, column=col, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = bg
            cell.border    = thin_border()
            cell.alignment = Alignment(horizontal="center")

    # Totales
    total_fila = len(resumen) + 2
    total_bytes = sum(r["tamanio_bytes"] for r in registros)
    for col, val in enumerate(["TOTAL", len(registros), formato_tamanio(total_bytes)], 1):
        cell = ws2.cell(row=total_fila, column=col, value=val)
        cell.font      = Font(name="Arial", bold=True, size=10)
        cell.fill      = PatternFill("solid", start_color="FFE2EFDA")
        cell.border    = thin_border()
        cell.alignment = Alignment(horizontal="center")

    # ── Metadata ─────────────────────────────────────────────────────────────
    wb.properties.title   = "Índice de biblioteca"
    wb.properties.creator = "index_libros.py"

    wb.save(output_path)
    return len(registros), resumen


def main():
    parser = argparse.ArgumentParser(description="Indexa libros en carpetas y genera un Excel.")
    parser.add_argument("carpeta", help="Carpeta raíz a escanear")
    parser.add_argument("--output", "-o", default="", help="Nombre del archivo de salida (default: biblioteca_YYYYMMDD.xlsx)")
    args = parser.parse_args()

    raiz = Path(args.carpeta).resolve()
    if not raiz.is_dir():
        print(f"Error: '{raiz}' no es una carpeta válida.")
        sys.exit(1)

    output = Path(args.output) if args.output else Path(f"biblioteca_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    print(f"Escaneando: {raiz}")
    registros = recorrer_carpeta(raiz)

    if not registros:
        print("No se encontraron archivos con las extensiones buscadas.")
        sys.exit(0)

    print(f"Archivos encontrados: {len(registros)}")
    total, resumen = crear_excel(registros, output, raiz)

    print(f"\nExcel generado: {output.resolve()}")
    print(f"{'Extensión':<12} {'Cantidad':>10} {'Tamaño':>12}")
    print("-" * 36)
    for ext, datos in sorted(resumen.items()):
        print(f"{ext:<12} {datos['cantidad']:>10} {formato_tamanio(datos['bytes']):>12}")
    print("-" * 36)
    print(f"{'TOTAL':<12} {total:>10} {formato_tamanio(sum(r['tamanio_bytes'] for r in registros)):>12}")


if __name__ == "__main__":
    main()